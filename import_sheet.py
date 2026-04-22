#!/usr/bin/env python3
"""
League of Thrones spreadsheet importer.

Reads the offseason roster xlsx (one tab per owner) and produces:
  - import_report.md   human-readable summary with anomalies flagged
  - seed_data.sql      SQL inserts for teams, contracts, draft picks

Run:
  python scripts/import_sheet.py path/to/sheet.xlsx --output-dir scripts/out

Nothing touches the database. Review the report, then apply the SQL via
Supabase SQL editor or psql.
"""
from __future__ import annotations

import argparse
import dataclasses
import re
import sys
import uuid
from pathlib import Path

from openpyxl import load_workbook


VALID_POSITIONS = {"QB", "RB", "WR", "TE", "DST", "K", "QB/TE", "QB/WR"}
ACQUISITION_SEASON = 2024  # Year-one prices in the offseason sheet were set in 2024
HIGH_PRICE_THRESHOLD = 80  # Flag (don't reject) prices over this for human review


@dataclasses.dataclass
class RosterEntry:
    player_name: str
    position: str | None
    contract_year: int | None
    year_one_price: int | None
    sheet_row: int


@dataclasses.dataclass
class DraftPickEntry:
    player_name: str
    cost: int | None
    sheet_row: int


@dataclasses.dataclass
class OwnerTab:
    owner: str
    roster: list[RosterEntry]
    draft_picks: list[DraftPickEntry]
    anomalies: list[str]


def find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """
    Locate the row containing 'Player' / 'Position' headers and return its
    index along with a column map. Handles tabs that are shifted right
    (Chav) or that have extra blank columns between headers (Chris).
    """
    target_headers = {
        "player": "player",
        "position": "position",
        "year of contract": "contract_year",
        "2024 acquisition price": "year_one_price",
    }
    for r_idx, row in enumerate(rows):
        cells = {
            c_idx: str(v).strip().lower()
            for c_idx, v in enumerate(row)
            if v is not None and str(v).strip()
        }
        if "player" not in cells.values():
            continue
        col_map: dict[str, int] = {}
        for c_idx, text in cells.items():
            if text in target_headers:
                col_map[target_headers[text]] = c_idx
        if "player" in col_map and "position" in col_map:
            return r_idx, col_map
    return None


def find_draft_picks_header(rows: list[tuple], start_row: int) -> tuple[int, int] | None:
    """
    Find the 'Draft Picks' section header and return (row_idx, col_idx_of_player_name).
    Per the sheet's layout, the cost column is two cells to the right of the name.
    """
    for r_idx in range(start_row, len(rows)):
        for c_idx, val in enumerate(rows[r_idx]):
            if val is not None and str(val).strip().lower() == "draft picks":
                return r_idx, c_idx
    return None


def coerce_int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name).strip()


def parse_owner_tab(sheet, owner: str) -> OwnerTab:
    rows = list(sheet.iter_rows(values_only=True))
    anomalies: list[str] = []
    roster: list[RosterEntry] = []
    picks: list[DraftPickEntry] = []

    header = find_header_row(rows)
    if header is None:
        anomalies.append("Could not find 'Player' / 'Position' header row.")
        return OwnerTab(owner, roster, picks, anomalies)

    header_row_idx, col_map = header
    col_player = col_map["player"]
    col_position = col_map["position"]
    col_year = col_map.get("contract_year")
    col_price = col_map.get("year_one_price")

    # Find the 'Draft Picks' / 'Offseason Transactions' divider so we know
    # where the roster section ends.
    end_row = len(rows)
    for r_idx in range(header_row_idx + 1, len(rows)):
        for c_val in rows[r_idx]:
            if isinstance(c_val, str):
                low = c_val.strip().lower()
                if low in ("draft picks", "offseason transactions"):
                    end_row = r_idx
                    break
        if end_row != len(rows):
            break

    # Walk roster rows.
    for r_idx in range(header_row_idx + 1, end_row):
        row = rows[r_idx]
        if col_player >= len(row):
            continue
        player_val = row[col_player]
        if player_val is None or not str(player_val).strip():
            continue

        name = normalize_name(str(player_val))
        position = (
            str(row[col_position]).strip()
            if col_position < len(row) and row[col_position]
            else None
        )
        contract_year = (
            coerce_int(row[col_year])
            if col_year is not None and col_year < len(row)
            else None
        )
        year_one_price = (
            coerce_int(row[col_price])
            if col_price is not None and col_price < len(row)
            else None
        )

        roster.append(RosterEntry(
            player_name=name,
            position=position,
            contract_year=contract_year,
            year_one_price=year_one_price,
            sheet_row=r_idx + 1,
        ))

        # Anomaly checks
        if contract_year is None or contract_year < 1 or contract_year > 4:
            anomalies.append(
                f"Row {r_idx + 1}: '{name}' has contract_year={contract_year} (must be 1-4)."
            )
        if year_one_price is None or year_one_price < 0:
            anomalies.append(
                f"Row {r_idx + 1}: '{name}' has year_one_price={year_one_price} (must be >= 0)."
            )
        elif year_one_price > HIGH_PRICE_THRESHOLD:
            anomalies.append(
                f"Row {r_idx + 1}: '{name}' has year_one_price=${year_one_price} "
                f"(over ${HIGH_PRICE_THRESHOLD}, please verify)."
            )
        if position and position.upper() not in VALID_POSITIONS:
            anomalies.append(
                f"Row {r_idx + 1}: '{name}' has unusual position '{position}'."
            )

    # Draft Picks section
    picks_header = find_draft_picks_header(rows, header_row_idx + 1)
    if picks_header is not None:
        picks_row_idx, picks_col_idx = picks_header
        cost_col_idx = picks_col_idx + 2  # 'Cost to Keep' is two cells right of name
        for r_idx in range(picks_row_idx + 1, len(rows)):
            row = rows[r_idx]
            if picks_col_idx >= len(row):
                continue
            pick_val = row[picks_col_idx]
            if isinstance(pick_val, str) and "total on the books" in pick_val.lower():
                break
            if pick_val is None or not str(pick_val).strip():
                continue
            name = normalize_name(str(pick_val))
            cost = coerce_int(row[cost_col_idx]) if cost_col_idx < len(row) else None
            picks.append(DraftPickEntry(
                player_name=name, cost=cost, sheet_row=r_idx + 1,
            ))
            if cost != 1:
                anomalies.append(
                    f"Row {r_idx + 1}: Draft pick entry '{name}' has cost=${cost} "
                    f"(rookie picks should be $1; will skip — likely owner scratch notes)."
                )

    return OwnerTab(owner, roster, picks, anomalies)


def slugify(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def sql_escape(s: str) -> str:
    return s.replace("'", "''")


def generate_sql(tabs: list[OwnerTab]) -> str:
    out: list[str] = [
        "-- Generated seed data from offseason xlsx import.",
        "-- Apply AFTER the schema migration (0001_initial_schema.sql).",
        "-- Wrapped in a transaction; any failure rolls back cleanly.",
        "",
        "BEGIN;",
        "",
        "-- Teams: one per owner. team.name is intentionally NULL — fill in via admin.",
    ]
    team_ids: dict[str, str] = {}
    for tab in tabs:
        team_id = str(uuid.uuid4())
        team_ids[tab.owner] = team_id
        out.append(
            "INSERT INTO teams (id, owner_name, slug) VALUES "
            f"('{team_id}', '{sql_escape(tab.owner)}', '{slugify(tab.owner)}');"
        )
    out.append("")

    out.append("-- Unmatched players: every roster name gets a placeholder row that the")
    out.append("-- contract references. Replaced with real Yahoo player IDs once OAuth is up.")
    unmatched_ids: dict[tuple[str, str, int], str] = {}  # (owner, name, sheet_row) -> uuid
    for tab in tabs:
        for entry in tab.roster:
            uid = str(uuid.uuid4())
            unmatched_ids[(tab.owner, entry.player_name, entry.sheet_row)] = uid
            position_sql = f"'{sql_escape(entry.position)}'" if entry.position else "NULL"
            out.append(
                "INSERT INTO unmatched_players (id, raw_name, position, source) VALUES "
                f"('{uid}', '{sql_escape(entry.player_name)}', {position_sql}, "
                f"'sheet_import_2025_offseason');"
            )
    out.append("")

    out.append("-- Contracts. acquisition_type is INFERRED from year_one_price:")
    out.append("--   $0 -> free_agent, $1 -> rookie_draft, anything else -> auction.")
    out.append("-- Waiver pickups (FAAB > $1) currently land as 'auction' and need to be")
    out.append("-- corrected via the admin panel. Known case: Isaac Guerendo ($100 FAAB).")
    skipped = 0
    for tab in tabs:
        team_id = team_ids[tab.owner]
        for entry in tab.roster:
            if entry.contract_year is None or entry.contract_year < 1 or entry.contract_year > 4:
                skipped += 1
                continue
            if entry.year_one_price is None or entry.year_one_price < 0:
                skipped += 1
                continue
            uid = unmatched_ids[(tab.owner, entry.player_name, entry.sheet_row)]
            if entry.year_one_price == 0:
                acq = "free_agent"
            elif entry.year_one_price == 1:
                acq = "rookie_draft"
            else:
                acq = "auction"
            out.append(
                "INSERT INTO contracts (unmatched_player_id, current_team_id, original_team_id, "
                "acquisition_season, acquisition_type, year_one_price, contract_year, status) VALUES "
                f"('{uid}', '{team_id}', '{team_id}', "
                f"{ACQUISITION_SEASON}, '{acq}', "
                f"{entry.year_one_price}, {entry.contract_year}, 'active');"
            )
    out.append("")
    if skipped:
        out.append(f"-- NOTE: {skipped} roster row(s) skipped due to invalid contract_year/price.")
        out.append("-- See import_report.md for details.")
        out.append("")

    out.append("-- Rookie draft picks for 2025 (only entries with cost == $1).")
    for tab in tabs:
        team_id = team_ids[tab.owner]
        for pick in tab.draft_picks:
            if pick.cost != 1:
                continue
            note = f"Pre-declared selection: {sql_escape(pick.player_name)}"
            out.append(
                "INSERT INTO draft_picks (season, original_team_id, current_team_id, "
                "pick_type, notes) VALUES "
                f"(2025, '{team_id}', '{team_id}', 'rookie', '{note}');"
            )
    out.append("")
    out.append("COMMIT;")
    out.append("")
    return "\n".join(out)


def generate_report(tabs: list[OwnerTab], xlsx_path: Path) -> str:
    out: list[str] = [
        "# Import Report",
        "",
        f"Source: `{xlsx_path.name}`",
        "",
        "## Summary",
        "",
    ]
    total_roster = sum(len(t.roster) for t in tabs)
    total_picks = sum(1 for t in tabs for p in t.draft_picks if p.cost == 1)
    total_anomalies = sum(len(t.anomalies) for t in tabs)
    out += [
        f"- Tabs parsed: **{len(tabs)}**",
        f"- Roster entries: **{total_roster}**",
        f"- Rookie draft picks at $1: **{total_picks}**",
        f"- Anomalies flagged: **{total_anomalies}**",
        "",
    ]
    if total_anomalies > 0:
        out.append("## Anomalies (review before applying SQL)")
        out.append("")
        for tab in tabs:
            if tab.anomalies:
                out.append(f"### {tab.owner}")
                out.append("")
                for a in tab.anomalies:
                    out.append(f"- {a}")
                out.append("")

    out.append("## Per-tab detail")
    out.append("")
    for tab in tabs:
        valid_picks = [p for p in tab.draft_picks if p.cost == 1]
        out.append(
            f"### {tab.owner}  ({len(tab.roster)} players, {len(valid_picks)} picks)"
        )
        out.append("")
        if tab.roster:
            out.append("| # | Player | Pos | Yr | Y1 $ |")
            out.append("|---|--------|-----|----|------|")
            for i, e in enumerate(tab.roster, 1):
                out.append(
                    f"| {i} | {e.player_name} | {e.position or '?'} | "
                    f"{e.contract_year if e.contract_year is not None else '?'} | "
                    f"${e.year_one_price if e.year_one_price is not None else '?'} |"
                )
            out.append("")
        if tab.draft_picks:
            out.append("**Draft picks section:**")
            out.append("")
            for p in tab.draft_picks:
                tag = "" if p.cost == 1 else " ⚠ skipped (not $1)"
                out.append(f"- {p.player_name} (${p.cost}){tag}")
            out.append("")
    return "\n".join(out)


def main() -> int:
    parser = argparse.ArgumentParser(description="Import LoT offseason xlsx.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("."))
    args = parser.parse_args()

    if not args.xlsx_path.exists():
        print(f"File not found: {args.xlsx_path}", file=sys.stderr)
        return 1

    args.output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(args.xlsx_path, read_only=True, data_only=True)
    tabs = [parse_owner_tab(wb[name], owner=name) for name in wb.sheetnames]

    report = generate_report(tabs, args.xlsx_path)
    sql = generate_sql(tabs)

    report_path = args.output_dir / "import_report.md"
    sql_path = args.output_dir / "seed_data.sql"
    report_path.write_text(report)
    sql_path.write_text(sql)

    print(f"Wrote {report_path}")
    print(f"Wrote {sql_path}")
    print(
        f"  {len(tabs)} tabs | "
        f"{sum(len(t.roster) for t in tabs)} roster rows | "
        f"{sum(len(t.anomalies) for t in tabs)} anomalies"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
